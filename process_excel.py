#!/usr/bin/env python3
"""处理普通食品 Excel：展开隐藏行列，合并最大允许限相关字段。"""

from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter

SOURCE = Path("/workspace/普通食品-模板_副本.xlsx")
OUTPUT_DIR = Path("/workspace/output")
ARTIFACTS_DIR = Path("/workspace/artifacts")
OUTPUT_NAME = "处理后数据.xlsx"
LIMIT_COL = "最大允许限"
UNIT_COL = "最大允许限单位"
NEW_COL = "数据单位"


def is_empty(value) -> bool:
    if pd.isna(value):
        return True
    text = str(value).strip()
    return text == "" or text.lower() == "nan"


def merge_data_unit(limit, unit) -> str:
    parts: list[str] = []
    if not is_empty(limit):
        parts.append(str(limit).strip())
    if not is_empty(unit):
        parts.append(str(unit).strip())
    return "".join(parts)


def unhide_sheet(ws) -> tuple[int, int]:
    hidden_rows = 0
    hidden_cols = 0

    for row_idx in range(1, ws.max_row + 1):
        dim = ws.row_dimensions[row_idx]
        if dim.hidden:
            dim.hidden = False
            hidden_rows += 1
        if dim.outlineLevel:
            dim.outlineLevel = 0

    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        dim = ws.column_dimensions[letter]
        if dim.hidden:
            dim.hidden = False
            hidden_cols += 1
        if dim.outlineLevel:
            dim.outlineLevel = 0

    if ws.sheet_properties and ws.sheet_properties.outlinePr:
        ws.sheet_properties.outlinePr.summaryBelow = True
        ws.sheet_properties.outlinePr.summaryRight = True

    return hidden_rows, hidden_cols


def process() -> dict:
    wb = openpyxl.load_workbook(SOURCE)
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    hidden_rows, hidden_cols = unhide_sheet(ws)

    df = pd.read_excel(SOURCE, sheet_name=sheet_name, header=0)
    source_rows = len(df)
    source_cols = len(df.columns)

    if LIMIT_COL not in df.columns or UNIT_COL not in df.columns:
        raise KeyError(f"缺少必要列：{LIMIT_COL} 或 {UNIT_COL}")

    limit_idx = df.columns.get_loc(LIMIT_COL)
    df[NEW_COL] = df.apply(
        lambda row: merge_data_unit(row[LIMIT_COL], row[UNIT_COL]),
        axis=1,
    )
    df.drop(columns=[LIMIT_COL, UNIT_COL], inplace=True)

    cols = list(df.columns)
    new_col = cols.pop(cols.index(NEW_COL))
    cols.insert(limit_idx, new_col)
    df = df[cols]

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    ARTIFACTS_DIR.mkdir(parents=True, exist_ok=True)

    output_path = OUTPUT_DIR / OUTPUT_NAME
    artifacts_path = ARTIFACTS_DIR / OUTPUT_NAME

    df.to_excel(output_path, sheet_name=sheet_name, index=False)
    shutil.copy2(output_path, artifacts_path)

    return {
        "sheet_name": sheet_name,
        "source_rows": source_rows,
        "source_cols": source_cols,
        "output_rows": len(df),
        "output_cols": len(df.columns),
        "hidden_rows_unhidden": hidden_rows,
        "hidden_cols_unhidden": hidden_cols,
        "output_path": str(output_path),
        "artifacts_path": str(artifacts_path),
    }


if __name__ == "__main__":
    result = process()
    print(result)
